"""
=============================================================================
EXCEL EXPORTER (THIN WRAPPER)
=============================================================================
Thin wrapper that calls engine modules for computation and formats results
to Excel.

Calls:
  - engine.eda_metrics.run_eda()
  - engine.modeling.run_modeling()
  - engine.response_curves.compute_response_curves()
  - engine.optimization_engine.run_optimization()
  
Formats output to multi-sheet Excel workbook.
"""

import pandas as pd
import numpy as np
from typing import Dict, Any, List, Optional
import warnings
warnings.filterwarnings('ignore')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# =============================================================================
# UTILITIES
# =============================================================================

def safe_write_df(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1):
    """Safely write DataFrame to Excel worksheet."""
    if df is None or df.empty:
        ws.cell(row=start_row, column=start_col).value = "(No data)"
        return
    
    # Headers
    for col_idx, col_name in enumerate(df.columns, start=start_col):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.value = str(col_name)
        if HAS_OPENPYXL:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Data
    for row_idx, (idx, row) in enumerate(df.iterrows(), start=start_row + 1):
        for col_idx, value in enumerate(row, start=start_col):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # Format numbers
            if isinstance(value, (int, float)):
                if isinstance(value, float) and value != int(value):
                    cell.number_format = '0.0000'
                else:
                    cell.number_format = '#,##0'
            
            cell.value = value
    
    # Auto-width (simplified)
    if HAS_OPENPYXL:
        for col_idx in range(start_col, start_col + len(df.columns)):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12


def safe_write_dict(ws, data_dict: Dict[str, Any], start_row: int = 1):
    """Write dict as key-value pairs to Excel."""
    row = start_row
    
    for key, value in data_dict.items():
        ws.cell(row=row, column=1).value = str(key)
        
        # Format value
        if isinstance(value, dict):
            ws.cell(row=row, column=2).value = str(value)
        elif isinstance(value, list):
            ws.cell(row=row, column=2).value = ", ".join(str(x) for x in value[:5])
        else:
            ws.cell(row=row, column=2).value = value
        
        row += 1


# =============================================================================
# EXPORT TO EXCEL
# =============================================================================

def export_to_excel(
    pipeline_results: Dict[str, Any],
    output_path: str
) -> Dict[str, Any]:
    """
    Export pipeline results to Excel workbook.
    
    Args:
        pipeline_results: Dict from orchestrator with all results
        output_path: Path to write .xlsx file
        
    Returns:
        {'success': bool, 'path': str, 'message': str}
    """
    if not HAS_OPENPYXL:
        return {
            'success': False,
            'message': 'openpyxl not installed. Install: pip install openpyxl'
        }
    
    try:
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        log = []
        
        # =====================================================================
        # SHEET 1: EDA SUMMARY
        # =====================================================================
        ws = wb.create_sheet("EDA Summary", 0)
        row = 1
        
        if 'eda_metrics' in pipeline_results:
            eda = pipeline_results['eda_metrics']
            
            ws.cell(row=row, column=1).value = "EDA Metrics"
            ws.cell(row=row, column=1).font = Font(bold=True, size=12)
            row += 2
            
            if 'eda_results' in eda and 'National' in eda['eda_results']:
                nat = eda['eda_results']['National']
                
                # Reach/Frequency/Engagement
                if 'rfe' in nat:
                    rfe_df = pd.DataFrame([nat['rfe']])
                    safe_write_df(ws, rfe_df, start_row=row)
                    row += len(rfe_df) + 3
                
                log.append("✅ EDA Summary sheet created")
        
        # =====================================================================
        # SHEET 2: MODELING RESULTS
        # =====================================================================
        ws = wb.create_sheet("Modeling", 1)
        row = 1
        
        if 'modeling' in pipeline_results:
            mod = pipeline_results['modeling']
            
            ws.cell(row=row, column=1).value = "Top Models"
            ws.cell(row=row, column=1).font = Font(bold=True, size=12)
            row += 2
            
            if 'ranked_models' in mod and mod['ranked_models']:
                # Convert ModelScore objects to dicts if needed
                models_list = []
                for m in mod['ranked_models'][:10]:
                    if hasattr(m, '__dict__'):
                        models_list.append(m.__dict__)
                    else:
                        models_list.append(m)
                
                models_df = pd.DataFrame(models_list)
                
                # Select display columns
                display_cols = [c for c in models_df.columns if c in 
                               ['model_type', 'r2', 'rmse', 'mae', 'cv_mean', 'feature_count']]
                
                if display_cols:
                    safe_write_df(ws, models_df[display_cols], start_row=row)
                    row += len(models_df) + 3
                
                log.append("✅ Modeling Results sheet created")
        
        # =====================================================================
        # SHEET 3: RESPONSE CURVES
        # =====================================================================
        ws = wb.create_sheet("Response Curves", 2)
        row = 1
        
        if 'response_curves' in pipeline_results:
            rc = pipeline_results['response_curves']
            
            ws.cell(row=row, column=1).value = "Response Curves"
            ws.cell(row=row, column=1).font = Font(bold=True, size=12)
            row += 2
            
            if 'curves' in rc:
                for channel, curve_data in rc['curves'].items():
                    ws.cell(row=row, column=1).value = f"Channel: {channel}"
                    ws.cell(row=row, column=1).font = Font(bold=True)
                    row += 1
                    
                    # Curve type and elasticity
                    ws.cell(row=row, column=1).value = "Curve Type"
                    ws.cell(row=row, column=2).value = curve_data.get('curve_type')
                    row += 1
                    
                    ws.cell(row=row, column=1).value = "Elasticity"
                    ws.cell(row=row, column=2).value = round(curve_data.get('elasticity', 0), 4)
                    row += 2
                
                log.append("✅ Response Curves sheet created")
        
        # =====================================================================
        # SHEET 4: OPTIMIZATION SCENARIOS
        # =====================================================================
        ws = wb.create_sheet("Optimization", 3)
        row = 1
        
        if 'optimization' in pipeline_results:
            opt = pipeline_results['optimization']
            
            ws.cell(row=row, column=1).value = "Optimization Scenarios"
            ws.cell(row=row, column=1).font = Font(bold=True, size=12)
            row += 2
            
            if 'scenarios' in opt:
                for scenario_name, scenario_data in opt['scenarios'].items():
                    if scenario_data is None or 'error' in scenario_data:
                        continue
                    
                    ws.cell(row=row, column=1).value = scenario_data.get('scenario', scenario_name)
                    ws.cell(row=row, column=1).font = Font(bold=True)
                    row += 1
                    
                    # Summary metrics
                    ws.cell(row=row, column=1).value = "Total Spend"
                    ws.cell(row=row, column=2).value = round(scenario_data.get('total_spend', 0), 0)
                    row += 1
                    
                    ws.cell(row=row, column=1).value = "Predicted Sales"
                    ws.cell(row=row, column=2).value = round(scenario_data.get('predicted_sales', 0), 0)
                    row += 1
                    
                    ws.cell(row=row, column=1).value = "Profit"
                    ws.cell(row=row, column=2).value = round(scenario_data.get('profit', 0), 0)
                    row += 1
                    
                    ws.cell(row=row, column=1).value = "ROI"
                    ws.cell(row=row, column=2).value = round(scenario_data.get('roi', 0), 3)
                    row += 2
                
                log.append("✅ Optimization Scenarios sheet created")
        
        # =====================================================================
        # SHEET 5: SUMMARY & METADATA
        # =====================================================================
        ws = wb.create_sheet("Summary", 4)
        row = 1
        
        ws.cell(row=row, column=1).value = "Pipeline Summary"
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 2
        
        # Mode and metadata
        if 'metadata' in pipeline_results:
            meta = pipeline_results['metadata']
            safe_write_dict(ws, {
                'Mode': meta.get('mode', 'N/A'),
                'Timestamp': meta.get('timestamp', 'N/A'),
                'Data Rows': meta.get('data_rows', 'N/A'),
                'Channels': meta.get('channels', 'N/A'),
            }, start_row=row)
        
        row += 10
        
        # All logs
        ws.cell(row=row, column=1).value = "Execution Logs"
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        all_logs = pipeline_results.get('log', [])
        for log_msg in all_logs:
            ws.cell(row=row, column=1).value = log_msg
            row += 1
        
        # Save
        wb.save(output_path)
        
        return {
            'success': True,
            'path': output_path,
            'message': f'Excel exported: {output_path}',
            'sheets_created': len(wb.sheetnames),
        }
    
    except Exception as e:
        return {
            'success': False,
            'path': output_path,
            'message': f'Export failed: {str(e)}',
        }
