"""
=============================================================================
EXCEL EXPORTER -- Export results to structured Excel workbook
=============================================================================
Creates multi-sheet Excel with:
  - Cleaned data
  - Feature matrix
  - Model rankings & coefficients
  - Response curves
  - Optimization scenarios
  - Summary narratives
=============================================================================
"""

import pandas as pd
import numpy as np
from typing import Dict, Any, List, Optional
from io import BytesIO
import warnings
warnings.filterwarnings('ignore')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# =============================================================================
# EXCEL FORMATTING
# =============================================================================

def get_header_style():
    """Excel header style."""
    if not HAS_OPENPYXL:
        return {}
    
    return {
        "fill": PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
        "font": Font(color="FFFFFF", bold=True),
        "alignment": Alignment(horizontal="center", vertical="center"),
    }


def write_dataframe_with_formatting(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1):
    """Write DataFrame to worksheet with basic formatting."""
    if not HAS_OPENPYXL:
        return
    
    # Headers
    for col_idx, col_name in enumerate(df.columns, start=start_col):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.value = col_name
        # Apply header style
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.font = Font(bold=True)
    
    # Data
    for row_idx, (idx, row) in enumerate(df.iterrows(), start=start_row + 1):
        for col_idx, value in enumerate(row, start=start_col):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # Format numbers
            if isinstance(value, float):
                if abs(value) < 1:
                    cell.number_format = '0.0000'
                else:
                    cell.number_format = '#,##0.00'
            
            cell.value = value
    
    # Auto-adjust column widths
    for col_idx in range(start_col, start_col + len(df.columns)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15


# =============================================================================
# SHEET BUILDERS
# =============================================================================

def create_sheet_eda_summary(
    wb,
    eda_results: Dict[str, Any]
) -> None:
    """Create EDA Summary sheet."""
    ws = wb.create_sheet("EDA Summary", 0)
    
    row = 1
    
    # National level
    ws[f'A{row}'] = "NATIONAL LEVEL"
    row += 1
    
    if "National" in eda_results:
        nat_data = eda_results["National"]
        
        # RFE table
        if "rfe" in nat_data:
            ws[f'A{row}'] = "Reach/Frequency/Engagement"
            row += 1
            
            rfe_data = nat_data["rfe"]
            rfe_df = pd.DataFrame([
                {
                    "Channel": channel,
                    "Reach %": metrics["reach_pct"],
                    "Frequency ($)": metrics["frequency_avg"],
                    "Engagement Corr": metrics["engagement_correlation"],
                    "Total Spend": metrics["total_spend"]
                }
                for channel, metrics in rfe_data.items()
            ])
            
            if HAS_OPENPYXL:
                write_dataframe_with_formatting(ws, rfe_df, row, 1)
            row += len(rfe_df) + 2
        
        # Correlations table
        if "correlations" in nat_data and "top_channels" in nat_data["correlations"]:
            ws[f'A{row}'] = "Top Channels by Correlation with Sales"
            row += 1
            
            top_channels = nat_data["correlations"]["top_channels"]
            corr_df = pd.DataFrame([
                {
                    "Rank": item["rank"],
                    "Channel": item["channel"],
                    "Correlation": item["correlation"]
                }
                for item in top_channels[:10]
            ])
            
            if HAS_OPENPYXL:
                write_dataframe_with_formatting(ws, corr_df, row, 1)
            row += len(corr_df) + 2
    
    # Segment level
    segment_names = [k for k in eda_results.keys() if k != "National"]
    if segment_names:
        row += 1
        ws[f'A{row}'] = "SEGMENT LEVEL SUMMARY"
        row += 1
        
        segment_summary_data = []
        for segment in segment_names:
            seg_data = eda_results[segment]
            segment_summary_data.append({
                "Segment": segment,
                "Channels Active": len([m for m in seg_data.get("rfe", {}).values() if m["reach_pct"] > 0]),
                "Top Correlation": seg_data.get("correlations", {}).get("top_channels", [{}])[0].get("correlation", 0)
            })
        
        segment_df = pd.DataFrame(segment_summary_data)
        if HAS_OPENPYXL:
            write_dataframe_with_formatting(ws, segment_df, row, 1)


def create_sheet_models(
    wb,
    ranked_models: List[Dict[str, Any]]
) -> None:
    """Create Models sheet with ranking and coefficients."""
    ws = wb.create_sheet("Models", 1)
    
    row = 1
    ws[f'A{row}'] = "TOP MODELS RANKING"
    row += 1
    
    # Ranking table
    ranking_data = []
    for model in ranked_models[:10]:
        ranking_data.append({
            "Rank": model.get("rank"),
            "Model Type": model.get("model_id"),
            "Composite Score": model.get("composite_score"),
            "Fit (R²)": model.get("fit_r2"),
            "Ordinality Score": model.get("ordinality"),
        })
    
    ranking_df = pd.DataFrame(ranking_data)
    if HAS_OPENPYXL:
        write_dataframe_with_formatting(ws, ranking_df, row, 1)
    
    row += len(ranking_df) + 3
    
    # Top model coefficients
    if ranked_models:
        top_model = ranked_models[0]
        if "model" in top_model:
            model_obj = top_model["model"]
            # Try to extract coefficients from sklearn model
            try:
                if hasattr(model_obj, "coef_"):
                    ws[f'A{row}'] = f"TOP MODEL ({top_model.get('type', 'Unknown')}) - COEFFICIENTS"
                    row += 1
                    
                    coef_values = model_obj.coef_
                    # Create feature names (assumed to be channel names from feature engineering)
                    channel_names = [f"Feature_{i}" for i in range(len(coef_values))]
                    
                    coef_data = []
                    for channel, coef in sorted(zip(channel_names, coef_values), key=lambda x: abs(x[1]), reverse=True):
                        coef_data.append({
                            "Channel": channel,
                            "Coefficient": round(float(coef), 4),
                            "Magnitude": abs(float(coef))
                        })
                    
                    if coef_data:
                        coef_df = pd.DataFrame(coef_data)
                        if HAS_OPENPYXL:
                            write_dataframe_with_formatting(ws, coef_df, row, 1)
            except Exception as e:
                # Skip coefficients if extraction fails
                pass


def create_sheet_response_curves(
    wb,
    elasticities: Dict[str, float]
) -> None:
    """Create Response Curves sheet with elasticities."""
    ws = wb.create_sheet("Response Curves", 2)
    
    row = 1
    ws[f'A{row}'] = "CHANNEL ELASTICITIES"
    row += 1
    
    elasticity_data = []
    for channel, elasticity in sorted(elasticities.items(), key=lambda x: abs(x[1]), reverse=True):
        elasticity_data.append({
            "Channel": channel,
            "Elasticity": round(elasticity, 4),
            "Interpretation": "Inelastic" if abs(elasticity) < 0.5 else "Elastic"
        })
    
    elasticity_df = pd.DataFrame(elasticity_data)
    if HAS_OPENPYXL:
        write_dataframe_with_formatting(ws, elasticity_df, row, 1)


def create_sheet_optimization(
    wb,
    scenarios: Dict[str, Any]
) -> None:
    """Create Optimization Scenarios sheet."""
    ws = wb.create_sheet("Optimization", 3)
    
    row = 1
    ws[f'A{row}'] = "OPTIMIZATION SCENARIOS COMPARISON"
    row += 2
    
    # Scenario overview
    scenario_summary = []
    for scenario_name, scenario_data in scenarios.items():
        scenario_summary.append({
            "Scenario": scenario_name,
            "Total Budget": round(scenario_data["total_budget"], 2),
            "Expected GMV": round(scenario_data["expected_gmv"], 2),
            "Expected ROI": round(scenario_data["expected_roi"], 4),
        })
    
    summary_df = pd.DataFrame(scenario_summary)
    if HAS_OPENPYXL:
        write_dataframe_with_formatting(ws, summary_df, row, 1)
    
    row += len(summary_df) + 3
    
    # Allocation details for each scenario
    for scenario_name, scenario_data in scenarios.items():
        ws[f'A{row}'] = f"{scenario_name.upper()} - ALLOCATION"
        row += 1
        
        allocation_data = []
        for channel, spend in scenario_data["allocation"].items():
            change = scenario_data["changes"].get(channel, 0)
            change_pct = (change / (spend - change)) * 100 if (spend - change) > 0 else 0
            
            allocation_data.append({
                "Channel": channel,
                "Budget": round(spend, 2),
                "Change ($)": round(change, 2),
                "Change (%)": round(change_pct, 1)
            })
        
        allocation_df = pd.DataFrame(allocation_data)
        if HAS_OPENPYXL:
            write_dataframe_with_formatting(ws, allocation_df, row, 1)
        
        row += len(allocation_df) + 2


def create_sheet_narratives(
    wb,
    narratives: Dict[str, str]
) -> None:
    """Create Narratives sheet with GenAI summaries."""
    ws = wb.create_sheet("Narratives", 4)
    
    ws.column_dimensions['A'].width = 100
    ws.column_dimensions['A'].alignment = Alignment(wrap_text=True, vertical="top")
    
    row = 1
    
    for narrative_name, narrative_text in narratives.items():
        ws[f'A{row}'] = f"{narrative_name.upper()}"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        ws[f'A{row}'] = narrative_text if narrative_text else "[No narrative generated]"
        ws.row_dimensions[row].height = 100  # Wrap text
        row += 3


# =============================================================================
# MAIN EXPORT FUNCTION
# =============================================================================

def export_to_excel(
    output_path: str,
    eda_results: Dict[str, Any] = None,
    ranked_models: List[Dict[str, Any]] = None,
    elasticities: Dict[str, float] = None,
    scenarios: Dict[str, Any] = None,
    narratives: Dict[str, str] = None,
) -> None:
    """
    Main entry point: export all results to Excel.
    
    Args:
        output_path: Path to save .xlsx file
        eda_results: EDA segment results
        ranked_models: Ranked model list
        elasticities: Channel elasticities
        scenarios: Optimization scenarios
        narratives: GenAI narratives
    """
    if not HAS_OPENPYXL:
        print("⚠️  openpyxl not installed. Cannot create Excel file.")
        print("    Install with: pip install openpyxl")
        return
    
    print(f"Creating Excel workbook: {output_path}")
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    if eda_results:
        create_sheet_eda_summary(wb, eda_results)
        print("✅ EDA Summary sheet created")
    
    if ranked_models:
        create_sheet_models(wb, ranked_models)
        print("✅ Models sheet created")
    
    if elasticities:
        create_sheet_response_curves(wb, elasticities)
        print("✅ Response Curves sheet created")
    
    if scenarios:
        create_sheet_optimization(wb, scenarios)
        print("✅ Optimization sheet created")
    
    if narratives:
        create_sheet_narratives(wb, narratives)
        print("✅ Narratives sheet created")
    
    # Save
    wb.save(output_path)
    print(f"\n✅ Excel file saved: {output_path}")
